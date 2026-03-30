#!/usr/bin/env python3
"""
Generate payslips as PDFs from the salary sheet Excel file.

Usage:
    python generate_payslips.py <excel_file> [options]

Options:
    --month TEXT          Month filter, e.g. "FEB 2026" (default: auto-detect from Attendance sheet)
    --employees TEXT      Comma-separated employee names to filter (default: all)
    --designation TEXT    Filter by designation, e.g. "PICKER / PACKER" or "SUPERVISOR"
    --output-dir TEXT     Output directory for PDFs (default: payslips_<month>/)
    --bw                  Generate black and white payslips (no colored backgrounds)
    --list                List all employees and designations, then exit

Examples:
    python generate_payslips.py "FEB_2026_ANAKAPALLI_TL4P SALARY SHEET.xlsx"
    python generate_payslips.py "FEB_2026_ANAKAPALLI_TL4P SALARY SHEET.xlsx" --designation "PICKER / PACKER"
    python generate_payslips.py "FEB_2026_ANAKAPALLI_TL4P SALARY SHEET.xlsx" --employees "C SUVARNA LAXMI,G LAVANYA"
    python generate_payslips.py "FEB_2026_ANAKAPALLI_TL4P SALARY SHEET.xlsx" --list
"""

import argparse
import hashlib
import io
import os
import sys
from datetime import datetime

import openpyxl
import qrcode
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.colors import HexColor, black, white
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT


def num_to_words(num):
    """Convert a number to words (Indian numbering style)."""
    ones = [
        "", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight",
        "Nine", "Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen",
        "Sixteen", "Seventeen", "Eighteen", "Nineteen",
    ]
    tens = [
        "", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy",
        "Eighty", "Ninety",
    ]

    if num == 0:
        return "Zero Rupees Only"

    num = int(round(num))
    if num < 0:
        return "Minus " + num_to_words(-num)

    def two_digits(n):
        if n < 20:
            return ones[n]
        return tens[n // 10] + (" " + ones[n % 10] if n % 10 else "")

    parts = []
    if num >= 10000000:
        parts.append(two_digits(num // 10000000) + " Crore")
        num %= 10000000
    if num >= 100000:
        parts.append(two_digits(num // 100000) + " Lakh")
        num %= 100000
    if num >= 1000:
        parts.append(two_digits(num // 1000) + " Thousand")
        num %= 1000
    if num >= 100:
        parts.append(ones[num // 100] + " Hundred")
        num %= 100
    if num > 0:
        parts.append(two_digits(num))

    return " ".join(parts) + " Rupees Only"


def fmt_amount(val):
    """Format amount with Indian comma style."""
    val = int(round(val))
    if val < 0:
        return "-" + fmt_amount(-val)
    s = str(val)
    if len(s) <= 3:
        return s
    result = s[-3:]
    s = s[:-3]
    while s:
        result = s[-2:] + "," + result
        s = s[:-2]
    return result


def read_wage_sheet(wb):
    """Read employee data from the Wage Sheet."""
    ws = wb["Wage Sheet"]
    employees = []

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=False):
        sno = row[0].value
        if sno is None or not isinstance(sno, (int, float)):
            continue

        emp = {
            "sno": int(sno),
            "emp_code": row[2].value,
            "name_attendance": row[3].value,
            "name_aadhar": row[4].value,
            "gender": row[5].value,
            "uan": row[6].value,
            "esic": row[7].value,
            "doj": row[8].value,
            "location": row[9].value,
            "designation": row[12].value,
            "month": row[14].value,
            "days_in_month": row[15].value,
            "working_days": row[16].value,
            "ot_hrs": row[17].value,
            "basic_monthly": row[18].value,
            "consolidated_actual": row[29].value,
            "hra_actual": row[30].value,
            "other_allow_actual": row[31].value,
            "conveyance_actual": row[32].value,
            "leave_actual": row[33].value,
            "ot_amount": row[34].value,
            "bonus_actual": row[35].value,
            "gross_actual": row[36].value,
            "pf_employee": row[45].value,
            "esic_employee": row[46].value,
            "lwf_employee": row[47].value,
            "pt": row[48].value,
            "total_deduction": row[49].value,
            "take_home": row[50].value,
        }
        employees.append(emp)

    return employees


def read_attendance(wb):
    """Read attendance data from the Attendence sheet."""
    ws = wb["Attendence"]
    month_cell = ws["D1"].value

    # Determine how many day columns exist (G onwards, up to col AH = index 33)
    # Day columns are G(6) through AH(33) = 28 max days
    attendance = {}
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=False):
        sno = row[0].value
        if sno is None or not isinstance(sno, (int, float)):
            continue

        emp_code = row[1].value

        # Count daily marks from columns G(6) to AH(33)
        counted_p = 0
        counted_pph = 0
        counted_wo = 0
        counted_a = 0
        for c in range(6, 34):
            val = row[c].value
            if val is None:
                continue
            mark = str(val).strip().upper()
            if mark == "PPH":
                counted_pph += 1
            elif mark == "P":
                counted_p += 1
            elif mark == "W/O":
                counted_wo += 1
            elif mark.startswith("A"):
                counted_a += 1

        attendance[emp_code] = {
            "name": row[2].value,
            "present": row[34].value or 0,       # AI - Present Days (sheet value)
            "absent": row[35].value or 0,         # AJ
            "week_offs": row[36].value or 0,      # AK
            "total_days": row[37].value or 0,     # AL
            "holidays": row[38].value or 0,       # AM
            "tllp": row[39].value or 0,           # AN
            # Counted from daily marks
            "counted_p": counted_p,
            "counted_pph": counted_pph,
            "counted_wo": counted_wo,
            "counted_a": counted_a,
            "counted_present": counted_p + counted_pph,  # PPH counts as present
        }

    return month_cell, attendance


def validate_data(employees, attendance, all_employees=None):
    """Validate attendance counts and cross-check with wage sheet. Returns list of warnings.

    employees: the filtered list to validate
    all_employees: the full list (used for the EXTRA check)
    """
    if all_employees is None:
        all_employees = employees
    warnings = []

    for emp in employees:
        code = emp["emp_code"]
        name = emp["name_attendance"] or emp["name_aadhar"] or f"Emp#{emp['sno']}"
        att = attendance.get(code)

        if att is None:
            warnings.append(f"[MISSING] {name} ({code}): Not found in Attendance sheet")
            continue

        prefix = f"{name} ({code})"

        # 1. Verify counted present days match sheet's Present Days (col AI)
        sheet_present = int(safe_num(att["present"]))
        counted_present = att["counted_present"]
        if counted_present != sheet_present:
            warnings.append(
                f"[PRESENT MISMATCH] {prefix}: "
                f"Counted {counted_present} present days (P={att['counted_p']}, PPH={att['counted_pph']}) "
                f"but Attendance sheet says {sheet_present}"
            )

        # 2. Verify counted week offs match sheet's Week Offs (col AK)
        sheet_wo = int(safe_num(att["week_offs"]))
        if att["counted_wo"] != sheet_wo:
            warnings.append(
                f"[WEEK OFF MISMATCH] {prefix}: "
                f"Counted {att['counted_wo']} week offs but Attendance sheet says {sheet_wo}"
            )

        # 3. Cross-check TLLP (Attendance AN) vs Working Days (Wage Sheet Q)
        wage_working_days = int(safe_num(emp["working_days"]))
        att_tllp = int(safe_num(att["tllp"]))
        if att_tllp != wage_working_days:
            warnings.append(
                f"[WORKING DAYS MISMATCH] {prefix}: "
                f"Attendance TLLP = {att_tllp} but Wage Sheet Working Days = {wage_working_days}"
            )

        # 4. Verify TLLP = Present + Public Holidays
        sheet_holidays = int(safe_num(att["holidays"]))
        expected_tllp = sheet_present + sheet_holidays
        if expected_tllp != att_tllp:
            warnings.append(
                f"[TLLP MISMATCH] {prefix}: "
                f"Present ({sheet_present}) + Holidays ({sheet_holidays}) = {expected_tllp} "
                f"but TLLP = {att_tllp}"
            )

        # 5. Verify total days adds up
        sheet_total = int(safe_num(att["total_days"]))
        expected_total = sheet_present + sheet_wo
        if expected_total != sheet_total:
            warnings.append(
                f"[TOTAL DAYS MISMATCH] {prefix}: "
                f"Present ({sheet_present}) + Week Offs ({sheet_wo}) = {expected_total} "
                f"but Total Days = {sheet_total}"
            )

    # Check for employees in Attendance but not in Wage Sheet (use full list)
    wage_codes = {e["emp_code"] for e in all_employees}
    for code, att in attendance.items():
        if code not in wage_codes:
            warnings.append(
                f"[EXTRA] {att['name']} ({code}): In Attendance sheet but not in Wage Sheet"
            )

    return warnings


def safe_num(val, default=0):
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def generate_qr_code(emp, month_str, net_pay, generated_at):
    """Generate a QR code image containing payslip verification data."""
    emp_name = (emp["name_aadhar"] or emp["name_attendance"] or "").strip()
    emp_code = emp["emp_code"] or ""
    uan = str(emp["uan"] or "")
    timestamp_str = generated_at.strftime("%d-%m-%Y %H:%M:%S")

    # Build a verification string and hash it (includes timestamp)
    verify_data = f"{emp_code}|{emp_name}|{month_str}|{int(round(net_pay))}|{timestamp_str}"
    verify_hash = hashlib.sha256(verify_data.encode()).hexdigest()[:12].upper()

    # QR content: readable verification info
    qr_content = (
        f"TECHNO SOLUTIONS - PAYSLIP\n"
        f"Employee: {emp_name}\n"
        f"Code: {emp_code}\n"
        f"UAN: {uan}\n"
        f"Month: {month_str}\n"
        f"Net Pay: {int(round(net_pay))}\n"
        f"Generated: {timestamp_str}\n"
        f"Verify: TS-{verify_hash}"
    )

    qr = qrcode.QRCode(version=1, box_size=4, border=1, error_correction=qrcode.constants.ERROR_CORRECT_M)
    qr.add_data(qr_content)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")

    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return buf, f"TS-{verify_hash}"


def generate_payslip_pdf(emp, attendance, month_date, output_path, serial_no, bw_mode=False):
    """Generate a single payslip PDF."""

    if isinstance(month_date, datetime):
        month_str = month_date.strftime("%B %Y").upper()
    else:
        month_str = str(month_date)

    att = attendance.get(emp["emp_code"], {})

    # Values
    basic = safe_num(emp["consolidated_actual"])
    hra = safe_num(emp["hra_actual"])
    conveyance = safe_num(emp["conveyance_actual"])
    bonus = safe_num(emp["bonus_actual"])
    leave = safe_num(emp["leave_actual"])
    ot_amount = safe_num(emp["ot_amount"])
    other_allow = safe_num(emp["other_allow_actual"])
    pf = safe_num(emp["pf_employee"])
    esic = safe_num(emp["esic_employee"])
    lwf = safe_num(emp["lwf_employee"])
    pt = safe_num(emp["pt"])
    gross = safe_num(emp["gross_actual"])
    total_ded = safe_num(emp["total_deduction"])
    net_pay = safe_num(emp["take_home"])
    days_in_month = int(safe_num(emp["days_in_month"]))
    working_days = int(safe_num(emp["working_days"]))

    # Colors - conditional based on bw_mode flag
    if bw_mode:
        header_bg = white
        light_blue = white
        light_green = white
        section_bg = white
        border_color = black
        header_text_color = black
        section_text_color = black
    else:
        header_bg = HexColor("#1F4E79")
        light_blue = HexColor("#D6E4F0")
        light_green = HexColor("#E2EFDA")
        section_bg = HexColor("#4472C4")
        border_color = HexColor("#7F7F7F")
        header_text_color = white
        section_text_color = white

    # Build PDF
    doc = SimpleDocTemplate(
        output_path,
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    elements = []

    # --- Header Table ---
    header_data = [
        ["TECHNO SOLUTIONS"],
        ["D.NO.6-12-36, 12/1 ARUNDEL PET, GUNTUR - 522002"],
        [f"Pay Slip for the Month of {month_str}"],
    ]
    header_table = Table(header_data, colWidths=[180 * mm])
    header_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), header_bg),
        ("TEXTCOLOR", (0, 0), (0, 0), header_text_color),
        ("FONTNAME", (0, 0), (0, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (0, 0), 18),
        ("ALIGN", (0, 0), (0, 2), "CENTER"),
        ("FONTNAME", (0, 1), (0, 1), "Helvetica"),
        ("FONTSIZE", (0, 1), (0, 1), 9),
        ("TEXTCOLOR", (0, 1), (0, 1), black),
        ("BACKGROUND", (0, 2), (0, 2), light_blue),
        ("FONTNAME", (0, 2), (0, 2), "Helvetica-Bold"),
        ("FONTSIZE", (0, 2), (0, 2), 11),
        ("BOX", (0, 0), (-1, -1), 0.5, border_color),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, border_color),
        ("TOPPADDING", (0, 0), (0, 0), 10),
        ("BOTTOMPADDING", (0, 0), (0, 0), 10),
        ("TOPPADDING", (0, 1), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
    ]))
    elements.append(header_table)

    # --- Generated date row ---
    gen_data = [
        ["", f"Pay Slip Generated Date: {datetime.now().strftime('%d-%m-%Y')}", "", f"#{serial_no}"],
    ]
    gen_table = Table(gen_data, colWidths=[20 * mm, 90 * mm, 50 * mm, 20 * mm])
    gen_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (3, 0), (3, 0), "RIGHT"),
        ("BOX", (0, 0), (-1, -1), 0.5, border_color),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    elements.append(gen_table)

    # --- Employee Details Table ---
    emp_name = emp["name_aadhar"] or emp["name_attendance"] or ""
    detail_data = [
        ["Employee Name", emp_name.strip(), "Location", emp["location"] or ""],
        ["Designation", emp["designation"] or "", "Month & Year", month_str],
        ["UAN NO", str(emp["uan"] or ""), "ESIC NO", str(emp["esic"] or "")],
        ["No Of Days This Month", str(days_in_month), "Days Worked", str(working_days)],
    ]
    detail_table = Table(detail_data, colWidths=[40 * mm, 55 * mm, 35 * mm, 50 * mm])
    detail_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, 0), (0, -1), light_blue),
        ("BACKGROUND", (2, 0), (2, -1), light_blue),
        ("BOX", (0, 0), (-1, -1), 0.5, border_color),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, border_color),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(detail_table)
    elements.append(Spacer(1, 3 * mm))

    # --- Earnings & Deductions Table ---
    earn_ded_data = [
        ["Earnings", "", "Deductions", ""],
    ]

    # Build earnings list
    earnings = [
        ("Basic - (A)", basic),
        ("HRA - (B)", hra),
        ("Conveyance - (C)", conveyance),
        ("BONUS - (D)", bonus),
        ("PL / Leave - (E)", leave),
    ]
    if other_allow > 0:
        earnings.append(("Other Allowance", other_allow))
    if ot_amount > 0:
        earnings.append(("OT Hours Amount", ot_amount))

    deductions = [
        ("PF (On Basic-A)", pf),
        ("E.S.I.C (0.75%)", esic),
        ("L W F", lwf),
        ("Professional Tax", pt),
    ]

    max_rows = max(len(earnings), len(deductions))
    for i in range(max_rows):
        e_label = earnings[i][0] if i < len(earnings) else ""
        e_val = fmt_amount(earnings[i][1]) if i < len(earnings) else ""
        d_label = deductions[i][0] if i < len(deductions) else ""
        d_val = fmt_amount(deductions[i][1]) if i < len(deductions) else ""
        earn_ded_data.append([e_label, e_val, d_label, d_val])

    # Totals row
    earn_ded_data.append(["Total Earnings", fmt_amount(gross), "Total Deduction", fmt_amount(total_ded)])
    # Net salary row
    earn_ded_data.append(["", "", "Net Salary", fmt_amount(net_pay)])

    ed_table = Table(earn_ded_data, colWidths=[50 * mm, 40 * mm, 50 * mm, 40 * mm])

    ed_style = [
        # Section headers
        ("BACKGROUND", (0, 0), (1, 0), section_bg),
        ("BACKGROUND", (2, 0), (3, 0), section_bg),
        ("TEXTCOLOR", (0, 0), (-1, 0), section_text_color),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        # Body
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        # Labels bold
        ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 1), (2, -1), "Helvetica-Bold"),
        # Amounts right-aligned
        ("ALIGN", (1, 1), (1, -1), "RIGHT"),
        ("ALIGN", (3, 1), (3, -1), "RIGHT"),
        # Totals row styling
        ("BACKGROUND", (0, -2), (-1, -2), light_blue),
        ("FONTNAME", (0, -2), (-1, -2), "Helvetica-Bold"),
        # Net salary row
        ("BACKGROUND", (2, -1), (3, -1), light_green),
        ("FONTNAME", (2, -1), (3, -1), "Helvetica-Bold"),
        ("FONTSIZE", (2, -1), (3, -1), 11),
        # Borders
        ("BOX", (0, 0), (-1, -1), 0.5, border_color),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, border_color),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (1, 0), (1, -1), 6),
        ("RIGHTPADDING", (3, 0), (3, -1), 6),
        # Vertical line between earnings and deductions
        ("LINEAFTER", (1, 0), (1, -1), 1, border_color),
    ]
    ed_table.setStyle(TableStyle(ed_style))
    elements.append(ed_table)
    elements.append(Spacer(1, 3 * mm))

    # --- Rupees in Words ---
    words_data = [
        ["Rupees In Words:", num_to_words(net_pay)],
    ]
    words_table = Table(words_data, colWidths=[35 * mm, 145 * mm])
    words_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, 0), "Helvetica-Bold"),
        ("FONTNAME", (1, 0), (1, 0), "Helvetica-Oblique"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOX", (0, 0), (-1, -1), 0.5, border_color),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(words_table)
    elements.append(Spacer(1, 6 * mm))

    # --- QR Code + Signature ---
    generated_at = datetime.now()
    qr_buf, verify_code = generate_qr_code(emp, month_str, net_pay, generated_at)
    qr_image = Image(qr_buf, width=28 * mm, height=28 * mm)

    # Build a nested table: QR on left, signature on right
    verify_style = ParagraphStyle(
        "verify", fontName="Courier", fontSize=7, leading=9,
        textColor=HexColor("#555555"),
    )
    timestamp_display = generated_at.strftime("%d-%m-%Y %H:%M:%S")
    verify_para = Paragraph(
        f"Verification: {verify_code}<br/>"
        f"Generated: {timestamp_display}<br/>"
        f"Auto-generated payslip",
        verify_style,
    )

    sig_inner = Table(
        [["For  TECHNO SOLUTIONS"], [""], ["Authorised Signatory"]],
        colWidths=[70 * mm],
    )
    sig_inner.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 2), (0, 2), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (0, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))

    footer_data = [[qr_image, verify_para, sig_inner]]
    footer_table = Table(footer_data, colWidths=[32 * mm, 75 * mm, 73 * mm])
    footer_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (0, 0), "LEFT"),
        ("ALIGN", (2, 0), (2, 0), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    elements.append(footer_table)

    doc.build(elements)


def main():
    parser = argparse.ArgumentParser(description="Generate payslip PDFs from salary sheet")
    parser.add_argument("excel_file", help="Path to the salary sheet Excel file")
    parser.add_argument("--month", help="Month filter, e.g. 'FEB 2026'")
    parser.add_argument("--employees", help="Comma-separated employee names")
    parser.add_argument("--designation", help="Filter by designation, e.g. 'PICKER / PACKER'")
    parser.add_argument("--output-dir", help="Output directory for PDFs")
    parser.add_argument("--bw", action="store_true", help="Generate black and white payslips (no colored backgrounds)")
    parser.add_argument("--list", action="store_true", help="List employees and exit")
    args = parser.parse_args()

    print(f"Loading workbook: {args.excel_file}")
    wb = openpyxl.load_workbook(args.excel_file, data_only=True)

    employees = read_wage_sheet(wb)
    month_date, attendance = read_attendance(wb)

    if not employees:
        print("No employee data found in Wage Sheet!")
        sys.exit(1)

    print(f"Found {len(employees)} employees in Wage Sheet")
    print(f"Found {len(attendance)} employees in Attendance sheet")

    if args.list:
        designations = set()
        print(f"\n{'No.':<5} {'Emp Code':<20} {'Name':<30} {'Designation':<20}")
        print("-" * 75)
        for emp in employees:
            name = emp["name_attendance"] or emp["name_aadhar"] or ""
            print(f"{emp['sno']:<5} {emp['emp_code'] or '':<20} {name:<30} {emp['designation']:<20}")
            designations.add(emp["designation"])
        print(f"\nDesignations: {', '.join(sorted(designations))}")
        return

    # Filter employees
    filtered = employees

    if args.employees:
        names = [n.strip().upper() for n in args.employees.split(",")]
        filtered = [
            e for e in filtered
            if (e["name_attendance"] and e["name_attendance"].strip().upper() in names)
            or (e["name_aadhar"] and e["name_aadhar"].strip().upper() in names)
        ]
        if not filtered:
            print(f"No employees matched: {args.employees}")
            print("Use --list to see all employee names")
            sys.exit(1)

    if args.designation:
        desig = args.designation.strip().upper()
        filtered = [
            e for e in filtered
            if e["designation"] and e["designation"].strip().upper() == desig
        ]
        if not filtered:
            print(f"No employees matched designation: {args.designation}")
            print("Use --list to see all designations")
            sys.exit(1)

    # Validate data
    print("\nValidating attendance vs wage sheet data...")
    warnings = validate_data(filtered, attendance, all_employees=employees)
    if warnings:
        print(f"\n{'='*60}")
        print(f"  {len(warnings)} WARNING(S) FOUND")
        print(f"{'='*60}")
        for w in warnings:
            print(f"  {w}")
        print(f"{'='*60}\n")

        response = input("Warnings found. Continue generating payslips? (y/n): ").strip().lower()
        if response != "y":
            print("Aborted.")
            sys.exit(0)
    else:
        print("All validations passed.\n")

    # Output directory
    if isinstance(month_date, datetime):
        month_str = month_date.strftime("%b_%Y").upper()
    else:
        month_str = "payslips"

    output_dir = args.output_dir or f"payslips_{month_str}"
    os.makedirs(output_dir, exist_ok=True)

    print(f"Generating {len(filtered)} payslip PDF(s) into {output_dir}/")

    for idx, emp in enumerate(filtered):
        name = (emp["name_attendance"] or emp["name_aadhar"] or f"employee_{emp['sno']}").strip()
        safe_name = name.replace(" ", "_").replace("/", "-").replace(".", "")
        filename = f"{safe_name}_{month_str}.pdf"
        filepath = os.path.join(output_dir, filename)

        generate_payslip_pdf(emp, attendance, month_date, filepath, idx + 1, args.bw)
        print(f"  [{idx + 1}/{len(filtered)}] {name} -> {filename}")

    print(f"\nDone! {len(filtered)} payslip(s) saved to {output_dir}/")


if __name__ == "__main__":
    main()
