#!/usr/bin/env python3
"""
GUI Payslip Generator using tkinter.
Reuses all core logic from generate_payslips.py.
"""

import os
import platform
import queue
import subprocess
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime

from generate_payslips import (
    read_wage_sheet,
    read_attendance,
    read_column_headers,
    validate_data,
    generate_payslip_pdf,
    load_config,
    save_config,
    safe_num,
    SYSTEM_COLUMNS,
    DEFAULT_CONFIG,
)

import openpyxl


class PayslipGeneratorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("TECHNO SOLUTIONS - Payslip Generator")
        self.root.minsize(750, 820)
        self.root.resizable(True, True)

        # State
        self.wb = None
        self.employees = []
        self.attendance = {}
        self.month_date = None
        self.designations = []
        self.generating = False
        self.msg_queue = queue.Queue()

        # Column state
        self.all_headers = []       # [(col_letter, header_name), ...]
        self.col_headers = {}       # col_letter -> header_name
        self.config = load_config()

        self._build_ui()
        self._poll_queue()

    def _col_display(self, col_letter):
        """Format column for display: 'Header Name (COL)'."""
        name = self.config.get("label_overrides", {}).get(col_letter)
        if not name:
            name = self.col_headers.get(col_letter, col_letter)
        return f"{name} ({col_letter})"

    def _col_letter_from_display(self, display_text):
        """Extract column letter from display text like 'Header Name (AD)'."""
        if "(" in display_text and display_text.endswith(")"):
            return display_text[display_text.rfind("(") + 1:-1]
        return display_text

    def _build_ui(self):
        pad = {"padx": 8, "pady": 4}

        # --- Excel file ---
        frame_file = ttk.LabelFrame(self.root, text="Excel File", padding=6)
        frame_file.pack(fill="x", **pad)

        self.var_filepath = tk.StringVar()
        ttk.Entry(frame_file, textvariable=self.var_filepath, width=60).pack(side="left", fill="x", expand=True, padx=(0, 4))
        ttk.Button(frame_file, text="Browse...", command=self._browse_file).pack(side="left")

        # --- Employee selection ---
        frame_emp = ttk.LabelFrame(self.root, text="Employees", padding=6)
        frame_emp.pack(fill="both", expand=True, **pad)

        self.var_selection = tk.StringVar(value="all")

        row_radios = ttk.Frame(frame_emp)
        row_radios.pack(fill="x")

        ttk.Radiobutton(row_radios, text="All Employees", variable=self.var_selection, value="all", command=self._on_selection_change).pack(side="left", padx=(0, 10))
        ttk.Radiobutton(row_radios, text="By Designation", variable=self.var_selection, value="designation", command=self._on_selection_change).pack(side="left", padx=(0, 4))

        self.combo_designation = ttk.Combobox(row_radios, state="disabled", width=25)
        self.combo_designation.pack(side="left", padx=(0, 10))

        ttk.Radiobutton(row_radios, text="Select Employees", variable=self.var_selection, value="select", command=self._on_selection_change).pack(side="left")

        # Listbox with scrollbar for employee selection
        list_frame = ttk.Frame(frame_emp)
        list_frame.pack(fill="both", expand=True, pady=(4, 0))

        scrollbar = ttk.Scrollbar(list_frame, orient="vertical")
        self.listbox = tk.Listbox(list_frame, selectmode="extended", yscrollcommand=scrollbar.set, height=8, state="disabled")
        scrollbar.config(command=self.listbox.yview)
        self.listbox.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="left", fill="y")

        # Select all / none buttons
        btn_frame = ttk.Frame(list_frame)
        btn_frame.pack(side="left", padx=(4, 0))
        self.btn_sel_all = ttk.Button(btn_frame, text="Select All", command=self._select_all_employees, state="disabled")
        self.btn_sel_all.pack(fill="x", pady=(0, 2))
        self.btn_sel_none = ttk.Button(btn_frame, text="Clear", command=self._clear_selection, state="disabled")
        self.btn_sel_none.pack(fill="x")

        # --- Payslip Columns ---
        self._build_column_selection(pad)

        # --- Options ---
        frame_opts = ttk.LabelFrame(self.root, text="Options", padding=6)
        frame_opts.pack(fill="x", **pad)

        self.var_bw = tk.BooleanVar(value=False)
        ttk.Checkbutton(frame_opts, text="Black && White (no colored backgrounds)", variable=self.var_bw).pack(anchor="w")

        self.var_show_days = tk.BooleanVar(value=self.config.get("show_days", True))
        ttk.Checkbutton(frame_opts, text="Show 'Days in Month' and 'Days Worked' in payslip", variable=self.var_show_days).pack(anchor="w")

        self.var_show_if_zero = tk.BooleanVar(value=self.config.get("show_if_zero", False))
        ttk.Checkbutton(frame_opts, text="Show earnings/deductions even when value is 0", variable=self.var_show_if_zero).pack(anchor="w")

        # --- Output dir ---
        frame_out = ttk.LabelFrame(self.root, text="Output Directory", padding=6)
        frame_out.pack(fill="x", **pad)

        self.var_outdir = tk.StringVar()
        ttk.Entry(frame_out, textvariable=self.var_outdir, width=60).pack(side="left", fill="x", expand=True, padx=(0, 4))
        ttk.Button(frame_out, text="Browse...", command=self._browse_outdir).pack(side="left")

        # --- Progress ---
        frame_prog = ttk.Frame(self.root, padding=(8, 4))
        frame_prog.pack(fill="x")

        self.progress = ttk.Progressbar(frame_prog, mode="determinate")
        self.progress.pack(side="left", fill="x", expand=True, padx=(0, 6))
        self.lbl_progress = ttk.Label(frame_prog, text="")
        self.lbl_progress.pack(side="left")

        # --- Log area ---
        frame_log = ttk.LabelFrame(self.root, text="Log / Warnings", padding=4)
        frame_log.pack(fill="both", expand=True, **pad)

        self.log_text = tk.Text(frame_log, height=6, state="disabled", wrap="word", font=("Courier", 10))
        log_scroll = ttk.Scrollbar(frame_log, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=log_scroll.set)
        self.log_text.pack(side="left", fill="both", expand=True)
        log_scroll.pack(side="left", fill="y")

        # Configure tag colors
        self.log_text.tag_configure("warning", foreground="#B8860B")
        self.log_text.tag_configure("error", foreground="#CC0000")
        self.log_text.tag_configure("success", foreground="#228B22")
        self.log_text.tag_configure("info", foreground="#333333")

        # --- Buttons ---
        frame_btns = ttk.Frame(self.root, padding=8)
        frame_btns.pack(fill="x")

        self.btn_generate = ttk.Button(frame_btns, text="Generate Payslips", command=self._start_generation, state="disabled")
        self.btn_generate.pack(side="left", padx=(0, 8))

        self.btn_save_defaults = ttk.Button(frame_btns, text="Save as Defaults", command=self._save_defaults, state="disabled")
        self.btn_save_defaults.pack(side="left", padx=(0, 8))

        self.btn_open_folder = ttk.Button(frame_btns, text="Open Output Folder", command=self._open_output_folder, state="disabled")
        self.btn_open_folder.pack(side="left")

    def _build_column_selection(self, pad):
        """Build the 3-listbox column selection section."""
        frame_cols = ttk.LabelFrame(self.root, text="Payslip Columns", padding=6)
        frame_cols.pack(fill="both", expand=True, **pad)

        # Available columns (left)
        left_frame = ttk.Frame(frame_cols)
        left_frame.pack(side="left", fill="both", expand=True)

        ttk.Label(left_frame, text="Available Columns").pack()
        avail_list_frame = ttk.Frame(left_frame)
        avail_list_frame.pack(fill="both", expand=True)
        avail_scroll = ttk.Scrollbar(avail_list_frame, orient="vertical")
        self.list_available = tk.Listbox(avail_list_frame, selectmode="extended", yscrollcommand=avail_scroll.set, height=7, width=28)
        avail_scroll.config(command=self.list_available.yview)
        self.list_available.pack(side="left", fill="both", expand=True)
        avail_scroll.pack(side="left", fill="y")

        # Arrow buttons (available -> earnings/deductions)
        arrow_frame = ttk.Frame(frame_cols)
        arrow_frame.pack(side="left", padx=4, pady=4)
        ttk.Button(arrow_frame, text="-> Earn", width=8, command=self._move_to_earnings).pack(pady=2)
        ttk.Button(arrow_frame, text="-> Ded", width=8, command=self._move_to_deductions).pack(pady=2)

        # Earnings (middle)
        mid_frame = ttk.Frame(frame_cols)
        mid_frame.pack(side="left", fill="both", expand=True)

        ttk.Label(mid_frame, text="Earnings").pack()
        earn_list_frame = ttk.Frame(mid_frame)
        earn_list_frame.pack(fill="both", expand=True)
        earn_scroll = ttk.Scrollbar(earn_list_frame, orient="vertical")
        self.list_earnings = tk.Listbox(earn_list_frame, selectmode="extended", yscrollcommand=earn_scroll.set, height=7, width=28)
        earn_scroll.config(command=self.list_earnings.yview)
        self.list_earnings.pack(side="left", fill="both", expand=True)
        earn_scroll.pack(side="left", fill="y")

        earn_btn_frame = ttk.Frame(mid_frame)
        earn_btn_frame.pack(fill="x")
        ttk.Button(earn_btn_frame, text="Up", width=5, command=lambda: self._move_up(self.list_earnings)).pack(side="left", padx=2)
        ttk.Button(earn_btn_frame, text="Down", width=5, command=lambda: self._move_down(self.list_earnings)).pack(side="left", padx=2)
        ttk.Button(earn_btn_frame, text="Remove", width=7, command=lambda: self._remove_from_list(self.list_earnings)).pack(side="left", padx=2)

        # Deductions (right)
        right_frame = ttk.Frame(frame_cols)
        right_frame.pack(side="left", fill="both", expand=True)

        ttk.Label(right_frame, text="Deductions").pack()
        ded_list_frame = ttk.Frame(right_frame)
        ded_list_frame.pack(fill="both", expand=True)
        ded_scroll = ttk.Scrollbar(ded_list_frame, orient="vertical")
        self.list_deductions = tk.Listbox(ded_list_frame, selectmode="extended", yscrollcommand=ded_scroll.set, height=7, width=28)
        ded_scroll.config(command=self.list_deductions.yview)
        self.list_deductions.pack(side="left", fill="both", expand=True)
        ded_scroll.pack(side="left", fill="y")

        ded_btn_frame = ttk.Frame(right_frame)
        ded_btn_frame.pack(fill="x")
        ttk.Button(ded_btn_frame, text="Up", width=5, command=lambda: self._move_up(self.list_deductions)).pack(side="left", padx=2)
        ttk.Button(ded_btn_frame, text="Down", width=5, command=lambda: self._move_down(self.list_deductions)).pack(side="left", padx=2)
        ttk.Button(ded_btn_frame, text="Remove", width=7, command=lambda: self._remove_from_list(self.list_deductions)).pack(side="left", padx=2)

    # --- Column list manipulation ---

    def _move_to_earnings(self):
        self._move_selected(self.list_available, self.list_earnings)

    def _move_to_deductions(self):
        self._move_selected(self.list_available, self.list_deductions)

    def _move_selected(self, from_list, to_list):
        """Move selected items from one listbox to another."""
        sel = from_list.curselection()
        if not sel:
            return
        items = [from_list.get(i) for i in sel]
        for item in items:
            to_list.insert("end", item)
        # Remove in reverse order to preserve indices
        for i in reversed(sel):
            from_list.delete(i)

    def _remove_from_list(self, listbox):
        """Remove selected items from earnings/deductions and put back in available."""
        sel = listbox.curselection()
        if not sel:
            return
        items = [listbox.get(i) for i in sel]
        for item in items:
            self.list_available.insert("end", item)
        for i in reversed(sel):
            listbox.delete(i)

    def _move_up(self, listbox):
        """Move selected item up in the listbox."""
        sel = listbox.curselection()
        if not sel or sel[0] == 0:
            return
        for i in sel:
            text = listbox.get(i)
            listbox.delete(i)
            listbox.insert(i - 1, text)
            listbox.selection_set(i - 1)

    def _move_down(self, listbox):
        """Move selected item down in the listbox."""
        sel = listbox.curselection()
        if not sel or sel[-1] == listbox.size() - 1:
            return
        for i in reversed(sel):
            text = listbox.get(i)
            listbox.delete(i)
            listbox.insert(i + 1, text)
            listbox.selection_set(i + 1)

    def _get_listbox_col_letters(self, listbox):
        """Get list of column letters from a listbox."""
        letters = []
        for i in range(listbox.size()):
            text = listbox.get(i)
            letters.append(self._col_letter_from_display(text))
        return letters

    def _populate_column_lists(self):
        """Populate the Available/Earnings/Deductions listboxes from headers + config."""
        self.list_available.delete(0, "end")
        self.list_earnings.delete(0, "end")
        self.list_deductions.delete(0, "end")

        earnings_set = set(self.config.get("earnings", []))
        deductions_set = set(self.config.get("deductions", []))
        assigned = earnings_set | deductions_set

        # Populate earnings in config order
        for col in self.config.get("earnings", []):
            if col in self.col_headers:
                self.list_earnings.insert("end", self._col_display(col))

        # Populate deductions in config order
        for col in self.config.get("deductions", []):
            if col in self.col_headers:
                self.list_deductions.insert("end", self._col_display(col))

        # Populate available: all headers not in system columns and not assigned
        for col_letter, header_name in self.all_headers:
            if col_letter in SYSTEM_COLUMNS:
                continue
            if col_letter in assigned:
                continue
            self.list_available.insert("end", self._col_display(col_letter))

    # --- File loading ---

    def _browse_file(self):
        path = filedialog.askopenfilename(
            title="Select Salary Sheet Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
        )
        if path:
            self.var_filepath.set(path)
            self._load_workbook(path)

    def _load_workbook(self, path):
        self._clear_log()
        self._log(f"Loading workbook: {os.path.basename(path)}...", "info")

        try:
            self.wb = openpyxl.load_workbook(path, data_only=True)
        except Exception as e:
            self._log(f"Error loading workbook: {e}", "error")
            messagebox.showerror("Error", f"Failed to load workbook:\n{e}")
            return

        try:
            self.employees = read_wage_sheet(self.wb)
        except Exception as e:
            self._log(f"Error reading Wage Sheet: {e}", "error")
            messagebox.showerror("Error", f"Failed to read Wage Sheet:\n{e}")
            return

        try:
            self.month_date, self.attendance = read_attendance(self.wb)
        except Exception as e:
            self._log(f"Error reading Attendance sheet: {e}", "error")
            messagebox.showerror("Error", f"Failed to read Attendance sheet:\n{e}")
            return

        # Read column headers
        try:
            self.all_headers = read_column_headers(self.wb)
            self.col_headers = {letter: name for letter, name in self.all_headers}
        except Exception as e:
            self._log(f"Error reading column headers: {e}", "error")
            messagebox.showerror("Error", f"Failed to read column headers:\n{e}")
            return

        if not self.employees:
            self._log("No employee data found in Wage Sheet!", "error")
            messagebox.showwarning("Warning", "No employee data found in Wage Sheet.")
            return

        self._log(f"Found {len(self.employees)} employees in Wage Sheet", "success")
        self._log(f"Found {len(self.attendance)} entries in Attendance sheet", "info")
        self._log(f"Found {len(self.all_headers)} columns in header row", "info")

        # Populate designations
        desigs = sorted(set(e["designation"] for e in self.employees if e["designation"]))
        self.designations = desigs
        self.combo_designation["values"] = desigs
        if desigs:
            self.combo_designation.current(0)

        # Populate listbox
        self.listbox.config(state="normal")
        self.listbox.delete(0, "end")
        for emp in self.employees:
            name = (emp["name_attendance"] or emp["name_aadhar"] or f"Emp#{emp['sno']}").strip()
            desig = emp["designation"] or ""
            self.listbox.insert("end", f"{name}  [{desig}]")
        self.listbox.config(state="disabled")

        # Populate column selection lists
        self._populate_column_lists()

        # Set default output dir
        if isinstance(self.month_date, datetime):
            month_str = self.month_date.strftime("%b_%Y").upper()
        else:
            month_str = "payslips"
        default_out = os.path.join(os.path.dirname(path), f"payslips_{month_str}")
        self.var_outdir.set(default_out)

        # Enable buttons
        self.btn_generate.config(state="normal")
        self.btn_save_defaults.config(state="normal")
        self._on_selection_change()

    # --- Employee selection ---

    def _on_selection_change(self):
        mode = self.var_selection.get()
        if mode == "designation":
            self.combo_designation.config(state="readonly")
            self.listbox.config(state="disabled")
            self.btn_sel_all.config(state="disabled")
            self.btn_sel_none.config(state="disabled")
        elif mode == "select":
            self.combo_designation.config(state="disabled")
            self.listbox.config(state="normal")
            self.btn_sel_all.config(state="normal")
            self.btn_sel_none.config(state="normal")
        else:
            self.combo_designation.config(state="disabled")
            self.listbox.config(state="disabled")
            self.btn_sel_all.config(state="disabled")
            self.btn_sel_none.config(state="disabled")

    def _select_all_employees(self):
        self.listbox.select_set(0, "end")

    def _clear_selection(self):
        self.listbox.selection_clear(0, "end")

    def _get_filtered_employees(self):
        mode = self.var_selection.get()
        if mode == "all":
            return list(self.employees)
        elif mode == "designation":
            desig = self.combo_designation.get().strip().upper()
            return [e for e in self.employees if e["designation"] and e["designation"].strip().upper() == desig]
        elif mode == "select":
            indices = self.listbox.curselection()
            if not indices:
                return []
            return [self.employees[i] for i in indices]
        return []

    # --- Output dir ---

    def _browse_outdir(self):
        path = filedialog.askdirectory(title="Select Output Directory")
        if path:
            self.var_outdir.set(path)

    def _open_output_folder(self):
        outdir = self.var_outdir.get()
        if not outdir or not os.path.isdir(outdir):
            messagebox.showinfo("Info", "Output directory does not exist yet.")
            return
        system = platform.system()
        if system == "Darwin":
            subprocess.Popen(["open", outdir])
        elif system == "Windows":
            os.startfile(outdir)
        else:
            subprocess.Popen(["xdg-open", outdir])

    # --- Save defaults ---

    def _save_defaults(self):
        """Save current column selections as defaults to payslip_defaults.json."""
        self.config["earnings"] = self._get_listbox_col_letters(self.list_earnings)
        self.config["deductions"] = self._get_listbox_col_letters(self.list_deductions)
        self.config["show_days"] = self.var_show_days.get()
        self.config["show_if_zero"] = self.var_show_if_zero.get()

        try:
            save_config(self.config)
            self._log("Defaults saved to payslip_defaults.json", "success")
            messagebox.showinfo("Saved", "Column selections saved as defaults.")
        except Exception as e:
            self._log(f"Error saving defaults: {e}", "error")
            messagebox.showerror("Error", f"Failed to save defaults:\n{e}")

    # --- Generation ---

    def _start_generation(self):
        if self.generating:
            return
        if not self.employees:
            messagebox.showwarning("Warning", "No workbook loaded.")
            return

        filtered = self._get_filtered_employees()
        if not filtered:
            messagebox.showwarning("Warning", "No employees selected.")
            return

        earnings_cols = self._get_listbox_col_letters(self.list_earnings)
        deductions_cols = self._get_listbox_col_letters(self.list_deductions)

        if not earnings_cols:
            messagebox.showwarning("Warning", "No earnings columns selected.")
            return
        if not deductions_cols:
            messagebox.showwarning("Warning", "No deductions columns selected.")
            return

        outdir = self.var_outdir.get().strip()
        if not outdir:
            messagebox.showwarning("Warning", "Please specify an output directory.")
            return

        # Validate
        self._log("\nValidating attendance vs wage sheet data...", "info")
        warnings = validate_data(filtered, self.attendance, all_employees=self.employees)
        if warnings:
            for w in warnings:
                self._log(f"  {w}", "warning")
            proceed = messagebox.askyesno(
                "Validation Warnings",
                f"{len(warnings)} warning(s) found.\n\nSee the log for details.\n\nContinue generating payslips?",
            )
            if not proceed:
                self._log("Generation aborted by user.", "info")
                return
        else:
            self._log("All validations passed.", "success")

        # Start generation in background thread
        self.generating = True
        self.btn_generate.config(state="disabled")
        self.progress["value"] = 0
        self.progress["maximum"] = len(filtered)
        self.lbl_progress.config(text=f"0/{len(filtered)}")

        bw = self.var_bw.get()
        show_days = self.var_show_days.get()
        show_if_zero = self.var_show_if_zero.get()
        label_overrides = self.config.get("label_overrides", {})

        thread = threading.Thread(
            target=self._generate_thread,
            args=(filtered, outdir, bw, earnings_cols, deductions_cols, label_overrides, show_if_zero, show_days),
            daemon=True,
        )
        thread.start()

    def _generate_thread(self, filtered, outdir, bw, earnings_cols, deductions_cols, label_overrides, show_if_zero, show_days):
        try:
            os.makedirs(outdir, exist_ok=True)

            if isinstance(self.month_date, datetime):
                month_str = self.month_date.strftime("%b_%Y").upper()
            else:
                month_str = "payslips"

            for idx, emp in enumerate(filtered):
                name = (emp["name_attendance"] or emp["name_aadhar"] or f"employee_{emp['sno']}").strip()
                safe_name = name.replace(" ", "_").replace("/", "-").replace(".", "")
                filename = f"{safe_name}_{month_str}.pdf"
                filepath = os.path.join(outdir, filename)

                generate_payslip_pdf(
                    emp, self.attendance, self.month_date, filepath, idx + 1, bw,
                    earnings_cols=earnings_cols,
                    deductions_cols=deductions_cols,
                    col_headers=self.col_headers,
                    label_overrides=label_overrides,
                    show_if_zero=show_if_zero,
                    show_days=show_days,
                )

                self.msg_queue.put(("progress", idx + 1, len(filtered), f"Generated: {filename}"))

            self.msg_queue.put(("done", len(filtered), outdir))
        except Exception as e:
            self.msg_queue.put(("error", str(e)))

    def _poll_queue(self):
        try:
            while True:
                msg = self.msg_queue.get_nowait()
                kind = msg[0]
                if kind == "progress":
                    _, current, total, text = msg
                    self.progress["value"] = current
                    self.lbl_progress.config(text=f"{current}/{total}")
                    self._log(f"  [{current}/{total}] {text}", "info")
                elif kind == "done":
                    _, count, outdir = msg
                    self._log(f"\nDone! {count} payslip(s) saved to {outdir}/", "success")
                    self.generating = False
                    self.btn_generate.config(state="normal")
                    self.btn_open_folder.config(state="normal")
                    messagebox.showinfo("Complete", f"{count} payslip(s) generated successfully!")
                elif kind == "error":
                    _, err = msg
                    self._log(f"\nError during generation: {err}", "error")
                    self.generating = False
                    self.btn_generate.config(state="normal")
                    messagebox.showerror("Error", f"Generation failed:\n{err}")
        except queue.Empty:
            pass
        self.root.after(100, self._poll_queue)

    # --- Log helpers ---

    def _log(self, text, tag="info"):
        self.log_text.config(state="normal")
        self.log_text.insert("end", text + "\n", tag)
        self.log_text.see("end")
        self.log_text.config(state="disabled")

    def _clear_log(self):
        self.log_text.config(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.config(state="disabled")


def main():
    root = tk.Tk()
    PayslipGeneratorApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
